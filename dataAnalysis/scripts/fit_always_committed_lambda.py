#!/usr/bin/env python3
"""Fit lambda for AlwaysCommittedAgent on pure human-human 2P3G data.

Model:
    W_lambda(g) ∝ exp(beta * EU(g)) * P_t(g)^lambda

Unlike the original committed-agent fitter, this scores all valid 2P3G action
steps because AlwaysCommittedAgent is active from trial start.
"""

from __future__ import annotations

import argparse
import ast
import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import minimize_scalar


GRID_SIZE = 15
GAMMA = 0.9
GOAL_REWARD = 30.0
STEP_COST = -1.0
SOFTMAX_BETA = 3.0
MODEL_BETA = 3.0
EPS = 1e-12
PROXIMITY_REWARD_WEIGHT = 0.01
PURE_PARTNER_TYPES = {"none", "human"}
ACTIONS: List[Tuple[int, int]] = [(0, -1), (0, 1), (-1, 0), (1, 0)]
ACTION_TO_INDEX = {action: idx for idx, action in enumerate(ACTIONS)}

FOLDERS = {
    "commitAgent-fallback": Path(
        "/Users/chengshaozhe/Documents/DukeECClab/code/collabAIdata/"
        "human-human-locaked action-with-commitAgent-fallback"
    ),
    "vlm-tom-fallback": Path(
        "/Users/chengshaozhe/Documents/DukeECClab/code/collabAIdata/"
        "human-human-locaked action-with-vlm-tom-fallback"
    ),
}

WORKBOOK_PARTICIPANT_FIELDS = {
    "participantId",
    "currentPlayer",
    "prolificPid",
    "trialStartTime",
    "endTime",
}
WORKBOOK_SUMMARY_FIELDS = {"totalSteps"}


@dataclass
class StepObservation:
    source_folder: str
    workbook_name: str
    room_id: str
    trial_index: int
    distance_condition: str
    step: int
    phase: str
    player: str
    self_position: Tuple[int, int]
    other_position: Tuple[int, int]
    goals: List[Tuple[int, int]]
    posterior: np.ndarray
    eu: np.ndarray
    action_goal_probs: np.ndarray
    observed_action: Tuple[int, int]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(
            "dataAnalysis/model_model/always_committed_agent/outputs/"
            "always_committed_lambda_fit_human_human"
        ),
    )
    parser.add_argument("--max-lambda", type=float, default=10.0)
    parser.add_argument("--beta", type=float, default=MODEL_BETA)
    return parser.parse_args()


def parse_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int)):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        lowered = text.lower()
        if lowered == "nan":
            return None
        if (
            text.startswith("[")
            or text.startswith("{")
            or "null" in lowered
            or "true" in lowered
            or "false" in lowered
        ):
            converted = (
                text.replace("null", "None")
                .replace("true", "True")
                .replace("false", "False")
            )
            try:
                return ast.literal_eval(converted)
            except Exception:
                return text
        try:
            number = float(text)
            return int(number) if number.is_integer() else number
        except ValueError:
            return text
    return value


def to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    if isinstance(value, str):
        try:
            number = float(value)
            return int(number) if number.is_integer() else None
        except ValueError:
            return None
    return None


def ensure_coord_list(value: Any) -> List[Tuple[int, int]]:
    parsed = parse_cell(value)
    if not isinstance(parsed, list):
        return []
    coords: List[Tuple[int, int]] = []
    for item in parsed:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            row = to_int(item[0])
            col = to_int(item[1])
            if row is not None and col is not None:
                coords.append((row, col))
    return coords


def ensure_action_list(value: Any) -> List[Tuple[int, int]]:
    return ensure_coord_list(value)


def workbook_rows(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["ExperimentData"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    out: List[Dict[str, Any]] = []
    for row in rows:
        if row is None or all(cell is None for cell in row):
            continue
        out.append({column: parse_cell(value) for column, value in zip(header, row)})
    return out


def is_pure_human_workbook(rows: Sequence[Dict[str, Any]]) -> bool:
    partner_types = {str(row.get("partnerAgentType") or "").strip() for row in rows}
    return "human" in partner_types and partner_types.issubset(PURE_PARTNER_TYPES)


def row_completeness_score(row: Dict[str, Any]) -> int:
    score = 0
    score += len(ensure_action_list(row.get("player1Actions")))
    score += len(ensure_action_list(row.get("player2Actions")))
    score += len(ensure_coord_list(row.get("player1Trajectory")))
    score += len(ensure_coord_list(row.get("player2Trajectory")))
    score += 5 if to_int(row.get("totalSteps")) not in (None, 0) else 0
    score += 5 if row.get("newGoalPresented") else 0
    score += 5 if row.get("newGoalPosition") is not None else 0
    return score


def canonical_core(row: Dict[str, Any]) -> Dict[str, Any]:
    core = {}
    for key, value in row.items():
        if key in WORKBOOK_PARTICIPANT_FIELDS or key in WORKBOOK_SUMMARY_FIELDS:
            continue
        core[key] = value
    return core


def transition(state: Tuple[int, int], move: Tuple[int, int]) -> Tuple[int, int]:
    nr = max(0, min(GRID_SIZE - 1, state[0] + move[0]))
    nc = max(0, min(GRID_SIZE - 1, state[1] + move[1]))
    return nr, nc


def action_probability(
    position: Tuple[int, int],
    other_position: Tuple[int, int],
    action: Tuple[int, int],
    goal: Tuple[int, int],
) -> float:
    if action not in ACTION_TO_INDEX:
        return EPS

    q_values = []
    for own_action in ACTIONS:
        own_next = position if position == goal else transition(position, own_action)
        for other_action in ACTIONS:
            other_next = other_position if other_position == goal else transition(other_position, other_action)
            done = own_next == goal and other_next == goal
            if done:
                reward = GOAL_REWARD
                future_value = 0.0
            else:
                joint_distance = (
                    abs(own_next[0] - goal[0]) + abs(own_next[1] - goal[1]) +
                    abs(other_next[0] - goal[0]) + abs(other_next[1] - goal[1])
                )
                reward = STEP_COST - PROXIMITY_REWARD_WEIGHT * joint_distance
                future_value = GAMMA * (GOAL_REWARD + STEP_COST * joint_distance)
            q_values.append(reward + future_value)

    q = np.array(q_values, dtype=np.float64)
    prefs = np.exp(np.clip(SOFTMAX_BETA * (q - np.max(q)), -700, 700))
    joint_probs = prefs / np.sum(prefs)
    own_idx = ACTION_TO_INDEX[action]
    return max(EPS, float(np.sum(joint_probs[own_idx * len(ACTIONS):(own_idx + 1) * len(ACTIONS)])))


def resize_posterior_like_always_committed(posterior: np.ndarray, goal_count: int) -> np.ndarray:
    if posterior.size == 0:
        return np.full(goal_count, 1.0 / goal_count, dtype=np.float64)
    if posterior.size == goal_count:
        return posterior

    next_post = np.zeros(goal_count, dtype=np.float64)
    overlap = min(goal_count, posterior.size)
    old = posterior[:overlap]
    old_total = float(np.sum(old))
    added = max(0, goal_count - posterior.size)
    new_goal_mass = added / goal_count if added > 0 else 0.0
    old_mass = max(0.0, 1.0 - new_goal_mass)

    if overlap > 0:
        if old_total > 0:
            next_post[:overlap] = old / old_total * old_mass
        else:
            next_post[:overlap] = old_mass / overlap
    if added > 0:
        next_post[posterior.size:goal_count] = 1.0 / goal_count
    return normalize_posterior(next_post)


def normalize_posterior(posterior: np.ndarray) -> np.ndarray:
    total = float(np.sum(posterior))
    if not math.isfinite(total) or total <= 0:
        return np.full(posterior.shape, 1.0 / posterior.size, dtype=np.float64)
    return posterior / total


def eu_joint_distance(self_pos: Tuple[int, int], other_pos: Tuple[int, int], goal: Tuple[int, int]) -> float:
    self_dist = abs(self_pos[0] - goal[0]) + abs(self_pos[1] - goal[1])
    other_dist = abs(other_pos[0] - goal[0]) + abs(other_pos[1] - goal[1])
    return -float(self_dist + other_dist)


def row_goals_for_step(
    row: Dict[str, Any],
    step: int,
    initial_goals: Sequence[Tuple[int, int]],
    new_goal: Optional[Tuple[int, int]],
    new_goal_time: Optional[int],
) -> List[Tuple[int, int]]:
    goals = list(initial_goals)
    if row.get("newGoalPresented") and new_goal is not None and new_goal_time is not None and step >= new_goal_time:
        goals.append(new_goal)
    return goals


def row_to_observations(row: Dict[str, Any]) -> List[StepObservation]:
    initial_goals = ensure_coord_list(row.get("initialGoalPositions"))
    if not initial_goals:
        return []

    new_goal_coords = ensure_coord_list([row.get("newGoalPosition")] if row.get("newGoalPosition") else [])
    new_goal = new_goal_coords[0] if new_goal_coords else None
    new_goal_time = to_int(row.get("newGoalPresentedTime"))

    p1_actions = ensure_action_list(row.get("player1Actions"))
    p2_actions = ensure_action_list(row.get("player2Actions"))
    p1_traj = ensure_coord_list(row.get("player1Trajectory"))
    p2_traj = ensure_coord_list(row.get("player2Trajectory"))

    max_steps = max(len(p1_actions), len(p2_actions), len(p1_traj), len(p2_traj))
    posterior = np.full(len(initial_goals), 1.0 / len(initial_goals), dtype=np.float64)
    observations: List[StepObservation] = []

    for step in range(max_steps):
        goals = row_goals_for_step(row, step, initial_goals, new_goal, new_goal_time)
        posterior = resize_posterior_like_always_committed(posterior, len(goals))
        phase = "post_new_goal" if len(goals) > len(initial_goals) else "pre_new_goal"

        if step < len(p1_actions) and step < len(p1_traj) and step < len(p2_traj):
            p1_pos = p1_traj[step]
            p2_pos = p2_traj[step]
            action_probs = np.array(
                [action_probability(p1_pos, p2_pos, p1_actions[step], goal) for goal in goals],
                dtype=np.float64,
            )
            observations.append(
                StepObservation(
                    source_folder=str(row["source_folder"]),
                    workbook_name=str(row["workbook_name"]),
                    room_id=str(row["roomId"]),
                    trial_index=to_int(row["trialIndex"]) or 0,
                    distance_condition=str(row.get("distanceCondition") or "unknown"),
                    step=step,
                    phase=phase,
                    player="player1",
                    self_position=p1_pos,
                    other_position=p2_pos,
                    goals=goals,
                    posterior=posterior.copy(),
                    eu=np.array([eu_joint_distance(p1_pos, p2_pos, goal) for goal in goals], dtype=np.float64),
                    action_goal_probs=action_probs,
                    observed_action=p1_actions[step],
                )
            )

        if step < len(p2_actions) and step < len(p2_traj) and step < len(p1_traj):
            p2_pos = p2_traj[step]
            p1_pos = p1_traj[step]
            action_probs = np.array(
                [action_probability(p2_pos, p1_pos, p2_actions[step], goal) for goal in goals],
                dtype=np.float64,
            )
            observations.append(
                StepObservation(
                    source_folder=str(row["source_folder"]),
                    workbook_name=str(row["workbook_name"]),
                    room_id=str(row["roomId"]),
                    trial_index=to_int(row["trialIndex"]) or 0,
                    distance_condition=str(row.get("distanceCondition") or "unknown"),
                    step=step,
                    phase=phase,
                    player="player2",
                    self_position=p2_pos,
                    other_position=p1_pos,
                    goals=goals,
                    posterior=posterior.copy(),
                    eu=np.array([eu_joint_distance(p2_pos, p1_pos, goal) for goal in goals], dtype=np.float64),
                    action_goal_probs=action_probs,
                    observed_action=p2_actions[step],
                )
            )

        like = np.ones(len(goals), dtype=np.float64)
        if step < len(p1_actions) and step < len(p1_traj) and step < len(p2_traj):
            like *= np.array(
                [action_probability(p1_traj[step], p2_traj[step], p1_actions[step], goal) for goal in goals],
                dtype=np.float64,
            )
        if step < len(p2_actions) and step < len(p2_traj) and step < len(p1_traj):
            like *= np.array(
                [action_probability(p2_traj[step], p1_traj[step], p2_actions[step], goal) for goal in goals],
                dtype=np.float64,
            )
        posterior = normalize_posterior(posterior * like)

    return observations


def lambda_negative_log_likelihood(lambda_value: float, observations: Sequence[StepObservation], beta: float) -> float:
    nll = 0.0
    for obs in observations:
        log_weights = beta * obs.eu + lambda_value * np.log(np.maximum(obs.posterior, EPS))
        max_log = float(np.max(log_weights))
        weights = np.exp(np.clip(log_weights - max_log, -700, 700))
        weights /= np.sum(weights)
        mixture_prob = float(np.dot(weights, obs.action_goal_probs))
        nll -= math.log(max(EPS, mixture_prob))
    return nll


def approximate_standard_error(lambda_value: float, observations: Sequence[StepObservation], beta: float) -> Optional[float]:
    if lambda_value <= 0:
        return None
    step = max(1e-3, lambda_value * 1e-3)
    left = max(0.0, lambda_value - step)
    center = lambda_negative_log_likelihood(lambda_value, observations, beta)
    right = lambda_negative_log_likelihood(lambda_value + step, observations, beta)
    if left == lambda_value:
        return None
    left_val = lambda_negative_log_likelihood(left, observations, beta)
    hessian = (right - 2.0 * center + left_val) / ((lambda_value - left) * step)
    if not math.isfinite(hessian) or hessian <= 0:
        return None
    return math.sqrt(1.0 / hessian)


def load_pure_human_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    pure_workbooks: List[Dict[str, Any]] = []
    pure_rows_2p3g: List[Dict[str, Any]] = []
    duplicate_rows: List[Dict[str, Any]] = []

    for source_label, folder in FOLDERS.items():
        for workbook in sorted(folder.glob("*.xlsx")):
            rows = workbook_rows(workbook)
            if not is_pure_human_workbook(rows):
                continue
            partner_types = sorted({str(row.get("partnerAgentType") or "") for row in rows})
            pure_workbooks.append(
                {
                    "source_folder": source_label,
                    "workbook_name": workbook.name,
                    "roomId": next((row.get("roomId") for row in rows if row.get("roomId")), None),
                    "participantId": next((row.get("participantId") for row in rows if row.get("participantId")), None),
                    "partnerAgentTypes": json.dumps(partner_types),
                    "human_2P3G_rows": sum(
                        1
                        for row in rows
                        if row.get("experimentType") == "2P3G" and row.get("partnerAgentType") == "human"
                    ),
                }
            )
            for row in rows:
                if row.get("experimentType") == "2P3G" and row.get("partnerAgentType") == "human":
                    enriched = dict(row)
                    enriched["source_folder"] = source_label
                    enriched["workbook_name"] = workbook.name
                    pure_rows_2p3g.append(enriched)

    grouped: Dict[Tuple[str, int], List[Dict[str, Any]]] = {}
    for row in pure_rows_2p3g:
        room_id = str(row.get("roomId"))
        trial_index = to_int(row.get("trialIndex"))
        if room_id and trial_index is not None:
            grouped.setdefault((room_id, trial_index), []).append(row)

    deduped_rows: List[Dict[str, Any]] = []
    for (room_id, trial_index), rows in grouped.items():
        chosen = max(rows, key=row_completeness_score)
        deduped_rows.append(chosen)

        core_reference = canonical_core(chosen)
        differing_columns = set()
        for other in rows:
            core_other = canonical_core(other)
            for key, value in core_reference.items():
                if core_other.get(key) != value:
                    differing_columns.add(key)
        duplicate_rows.append(
            {
                "roomId": room_id,
                "trialIndex": trial_index,
                "duplicate_count": len(rows),
                "chosen_workbook": chosen["workbook_name"],
                "source_folder": chosen["source_folder"],
                "differing_columns": json.dumps(sorted(differing_columns)),
            }
        )

    return pure_workbooks, pure_rows_2p3g, deduped_rows, duplicate_rows


def build_lambda_profile(observations: Sequence[StepObservation], beta: float, max_lambda: float) -> pd.DataFrame:
    grid = sorted(set(
        [round(i / 10, 1) for i in range(11)] +
        [round(i / 2, 1) for i in range(0, int(max_lambda * 2) + 1)]
    ))
    rows = []
    for value in grid:
        if value <= max_lambda + 1e-9:
            nll = lambda_negative_log_likelihood(value, observations, beta)
            rows.append({
                "lambda": value,
                "negative_log_likelihood": nll,
                "log_likelihood": -nll,
            })
    return pd.DataFrame(rows)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    pure_workbooks, pure_rows_2p3g, deduped_rows, duplicate_rows = load_pure_human_rows()

    observations: List[StepObservation] = []
    for row in deduped_rows:
        observations.extend(row_to_observations(row))

    if not observations:
        raise ValueError("No step-level observations found for AlwaysCommittedAgent fit.")

    result = minimize_scalar(
        lambda value: lambda_negative_log_likelihood(value, observations, args.beta),
        bounds=(0.0, float(args.max_lambda)),
        method="bounded",
    )

    lambda_hat = float(result.x)
    neg_log_likelihood = float(result.fun)
    log_likelihood = -neg_log_likelihood
    se = approximate_standard_error(lambda_hat, observations, args.beta)
    ci_low = max(0.0, lambda_hat - 1.96 * se) if se is not None else None
    ci_high = lambda_hat + 1.96 * se if se is not None else None

    workbooks_df = pd.DataFrame(pure_workbooks).sort_values(["source_folder", "roomId", "participantId"])
    deduped_df = pd.DataFrame(
        [
            {
                "source_folder": row["source_folder"],
                "workbook_name": row["workbook_name"],
                "roomId": row["roomId"],
                "trialIndex": to_int(row["trialIndex"]),
                "distanceCondition": row.get("distanceCondition"),
                "newGoalPresented": bool(row.get("newGoalPresented")),
                "newGoalPresentedTime": to_int(row.get("newGoalPresentedTime")),
                "firstDetectedSharedGoal": to_int(row.get("firstDetectedSharedGoal")),
                "partnerAgentType": row.get("partnerAgentType"),
            }
            for row in deduped_rows
        ]
    ).sort_values(["roomId", "trialIndex"])
    duplicates_df = pd.DataFrame(duplicate_rows).sort_values(["roomId", "trialIndex"])
    observations_df = pd.DataFrame(
        [
            {
                "source_folder": obs.source_folder,
                "workbook_name": obs.workbook_name,
                "roomId": obs.room_id,
                "trialIndex": obs.trial_index,
                "distanceCondition": obs.distance_condition,
                "step": obs.step,
                "phase": obs.phase,
                "player": obs.player,
                "selfPosition": json.dumps(list(obs.self_position)),
                "otherPosition": json.dumps(list(obs.other_position)),
                "goals": json.dumps([list(goal) for goal in obs.goals]),
                "posterior": json.dumps(obs.posterior.tolist()),
                "eu": json.dumps(obs.eu.tolist()),
                "actionGoalProbs": json.dumps(obs.action_goal_probs.tolist()),
                "observedAction": json.dumps(list(obs.observed_action)),
            }
            for obs in observations
        ]
    ).sort_values(["roomId", "trialIndex", "step", "player"])
    profile_df = build_lambda_profile(observations, args.beta, args.max_lambda)

    phase_counts = observations_df["phase"].value_counts().to_dict()
    summary = {
        "model": "AlwaysCommittedAgent",
        "fit_target": "all_valid_2p3g_steps",
        "pure_human_workbooks": int(len(workbooks_df)),
        "pure_human_2p3g_participant_rows": int(len(pure_rows_2p3g)),
        "unique_2p3g_room_trials": int(len(deduped_rows)),
        "step_level_observations": int(len(observations)),
        "pre_new_goal_observations": int(phase_counts.get("pre_new_goal", 0)),
        "post_new_goal_observations": int(phase_counts.get("post_new_goal", 0)),
        "fitted_lambda": lambda_hat,
        "negative_log_likelihood": neg_log_likelihood,
        "log_likelihood": log_likelihood,
        "beta_fixed": float(args.beta),
        "std_error": se,
        "ci95_lower": ci_low,
        "ci95_upper": ci_high,
        "optimizer_success": bool(result.success),
        "optimizer_message": str(result.message),
        "posterior_resize": "new goals receive 1/current_goal_count mass; old posterior is rescaled preserving relative odds",
    }

    workbooks_df.to_csv(output_dir / "pure_human_workbooks.csv", index=False)
    deduped_df.to_csv(output_dir / "unique_room_trials_2p3g.csv", index=False)
    duplicates_df.to_csv(output_dir / "duplicate_validation_2p3g.csv", index=False)
    observations_df.to_csv(output_dir / "step_level_observations_2p3g.csv", index=False)
    profile_df.to_csv(output_dir / "lambda_profile.csv", index=False)
    (output_dir / "lambda_fit_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
